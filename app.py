import streamlit as st
import pandas as pd
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

st.set_page_config(
    page_title="USGEP Branş Yönlendirme",
    layout="wide"
)

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
ADMIN_PASSWORD = st.secrets["ADMIN_PASSWORD"]

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


def sporculari_getir():
    data = supabase.table("sporcular").select("*").order("id").execute()
    return pd.DataFrame(data.data)


def testleri_getir():
    data = supabase.table("testler").select("*").order("id").execute()
    return pd.DataFrame(data.data)


def temiz_deger_mi(deger):
    return deger not in [None, "", 0, 0.0]


RENKLER = {
    4: "FF0000",
    8: "FFC000",
    12: "5B9BD5",
    16: "A9D08E",
    20: "00B050",
}

TEST_SAYFA_ESLESME = {
    "BOY": "BOY",
    "KİLO": "KİLO",
    "KULAÇ": "KULAÇ",
    "DURARAK UZUN ATLAMA": "DURARAK UZUN ATLAMA",
    "DİKEY SIÇRAMA": "DİKEY SIÇRAMA",
    "EL KAVRAMA": "EL KAVRAMA",
    "GERİYE SAĞLIK TOPU ATMA": "GERİYE SAĞLIK TOPU FIRLATMA",
    "20 M. SPRINT": "20 M. SPRINT",
    "AYAK ÇABUKLUĞU": "AYAK ÇABUKLUĞU",
    "EL ÇABUKLUĞU": "EL ÇABUKLUĞU",
    "SIRT BACAK": "SIRT BACAK KUVVETİ",
    "HEXAGON": "HEXAGON",
    "LANE AGILITY": "LANE ÇEVİKLİK",
}

TESTLER = list(TEST_SAYFA_ESLESME.keys()) + ["FONKSİYONEL ÇÖMELME"]

CIKTI_TEST_SIRASI = [
    "BOY",
    "KİLO",
    "BACAK BOYU",
    "OTURMA YÜKSEKLİĞİ",
    "BACAK UZUNLUĞU",
    "KULAÇ",
    "FONKSİYONEL ÇÖMELME",
    "EL ÇABUKLUĞU",
    "AYAK ÇABUKLUĞU",
    "EL KAVRAMA",
    "SIRT BACAK",
    "HEXAGON",
    "DURARAK UZUN ATLAMA",
    "GERİYE SAĞLIK TOPU ATMA",
    "DİKEY SIÇRAMA",
    "LANE AGILITY",
    "20 M. SPRINT",
]

BRANS_KRITERLERI = {
    "KARATE": {
        "yas": [7, 8, 9],
        "cinsiyet": ["ERKEK", "KIZ"],
        "kriterler": [
            "BACAK UZUNLUĞU",
            "KULAÇ",
            "DURARAK UZUN ATLAMA",
            "AYAK ÇABUKLUĞU",
            "EL ÇABUKLUĞU",
        ],
    },
    "TEKVANDO": {
        "yas": [7, 8, 9],
        "cinsiyet": ["ERKEK", "KIZ"],
        "kriterler": [
            "BACAK UZUNLUĞU",
            "KULAÇ",
            "DURARAK UZUN ATLAMA",
            "AYAK ÇABUKLUĞU",
            "FONKSİYONEL ÇÖMELME",
        ],
    },
    "BOKS": {
        "yas": [9, 10, 11],
        "cinsiyet": ["ERKEK", "KIZ"],
        "kriterler": [
            "KULAÇ",
            "HEXAGON",
            "GERİYE SAĞLIK TOPU ATMA",
            "AYAK ÇABUKLUĞU",
            "EL ÇABUKLUĞU",
        ],
    },
    "JUDO": {
        "yas": [7, 8, 9, 10, 11],
        "cinsiyet": ["ERKEK", "KIZ"],
        "kriterler": [
            "DURARAK UZUN ATLAMA",
            "GERİYE SAĞLIK TOPU ATMA",
            "SIRT BACAK",
            "EL KAVRAMA",
            "AYAK ÇABUKLUĞU",
        ],
    },
    "GÜREŞ": {
        "yas": [9, 10],
        "cinsiyet": ["ERKEK"],
        "kriterler": [
            "DURARAK UZUN ATLAMA",
            "GERİYE SAĞLIK TOPU ATMA",
            "SIRT BACAK",
            "EL KAVRAMA",
            "FONKSİYONEL ÇÖMELME",
        ],
    },
}


def normalize_text(x):
    x = str(x).strip().upper()
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = x.replace("\n", " ")
    x = re.sub(r"\s+", " ", x)
    return x


def temizle_metin(x):
    return str(x).strip().upper()


def sayiya_cevir(x):
    return float(str(x).replace(",", ".").strip())


def kolon_bul(df, adaylar):
    normalized_cols = {normalize_text(c): c for c in df.columns}

    for aday in adaylar:
        aday_norm = normalize_text(aday)

        for norm_col, original_col in normalized_cols.items():
            if aday_norm in norm_col:
                return original_col

    return None


def aralik_uyuyor_mu(deger, aralik):
    metin = str(aralik).replace(",", ".").strip()
    sayilar = re.findall(r"\d+(?:\.\d+)?", metin)

    if not sayilar:
        return False

    sayilar = [float(s) for s in sayilar]

    if "≤" in metin or "<=" in metin:
        return deger <= sayilar[0]

    if "≥" in metin or ">=" in metin:
        return deger >= sayilar[0]

    if "-" in metin and len(sayilar) >= 2:
        alt = min(sayilar[0], sayilar[1])
        ust = max(sayilar[0], sayilar[1])
        return alt <= deger <= ust

    return False


def fonksiyonel_puanla(deger):
    deger = str(deger).strip()

    if deger == "1":
        return 4
    elif deger == "2":
        return 8
    elif deger == "3":
        return 16
    elif deger == "4":
        return 20

    return 4


def norm_puanla(deger, norm_satiri):
    if pd.isna(deger) or temizle_metin(deger) in ["VERİ YOK", "VERI YOK", "G", "K", "NAN", ""]:
        return 4

    try:
        deger = sayiya_cevir(deger)
    except Exception:
        return 4

    kolon_puanlari = {
        "ÇOK ALTI": 4,
        "ALTI": 8,
        "ORTALAMA": 12,
        "ÜSTÜ": 16,
        "ÇOK ÜSTÜ": 20,
    }

    for kolon, puan in kolon_puanlari.items():
        if kolon in norm_satiri and aralik_uyuyor_mu(deger, norm_satiri[kolon]):
            return puan

    return 4


def sheet_oku(norm_dosya, sheet_adi):
    excel = pd.ExcelFile(norm_dosya)
    sayfalar = {str(s).strip(): s for s in excel.sheet_names}

    if sheet_adi not in sayfalar:
        raise ValueError(
            f"Norm dosyasında '{sheet_adi}' sayfası bulunamadı. "
            f"Mevcut sayfalar: {excel.sheet_names}"
        )

    return pd.read_excel(norm_dosya, sheet_name=sayfalar[sheet_adi])


def yas_hesapla(ham):
    dogum_tarihi_col = kolon_bul(ham, ["DOĞUM TARİHİ", "DOGUM TARIHI"])
    dogum_yili_col = kolon_bul(ham, ["DOĞUM YILI", "DOGUM YILI"])

    if dogum_tarihi_col:
        ham["YAŞ"] = (
            pd.Timestamp.now().year
            - pd.to_datetime(ham[dogum_tarihi_col], errors="coerce").dt.year
        )
    elif dogum_yili_col:
        ham["YAŞ"] = (
            pd.Timestamp.now().year
            - pd.to_numeric(ham[dogum_yili_col], errors="coerce")
        )
    else:
        ham["YAŞ"] = 0

    ham["YAŞ"] = ham["YAŞ"].fillna(0).astype(int)
    return ham


def bacak_uzunlugu_hesapla(ham):
    bacak_col = kolon_bul(ham, ["BACAK BOYU", "BACAK BOY"])
    oturma_col = kolon_bul(
        ham,
        ["OTURMA YÜKSEKLİĞİ", "OTURMA YUKSEKLIGI", "OTURMA"]
    )

    if bacak_col and oturma_col:
        uzunluklar = []
        puanlar = []

        for _, row in ham.iterrows():
            try:
                bacak = sayiya_cevir(row[bacak_col])
                oturma = sayiya_cevir(row[oturma_col])

                if bacak > oturma * 0.90:
                    uzunluklar.append("UZUN BACAK")
                    puanlar.append(20)
                else:
                    uzunluklar.append("KISA BACAK")
                    puanlar.append(4)

            except Exception:
                uzunluklar.append("VERİ YOK")
                puanlar.append(4)

        ham["BACAK UZUNLUĞU"] = uzunluklar
        ham["BACAK UZUNLUĞU PUAN"] = puanlar

    else:
        ham["BACAK UZUNLUĞU"] = "VERİ YOK"
        ham["BACAK UZUNLUĞU PUAN"] = 4

    return ham


def islem_yap(ham, norm_dosya):
    ham = yas_hesapla(ham)
    ham = bacak_uzunlugu_hesapla(ham)

    for test in TESTLER:
        puan_sutunu = f"{test} PUAN"

        if test == "FONKSİYONEL ÇÖMELME":
            if test in ham.columns:
                ham[puan_sutunu] = ham[test].apply(fonksiyonel_puanla)
            else:
                ham[puan_sutunu] = 4
            continue

        if test not in ham.columns:
            ham[puan_sutunu] = 4
            continue

        sayfa = TEST_SAYFA_ESLESME[test]
        norm = sheet_oku(norm_dosya, sayfa)

        puanlar = []

        for _, row in ham.iterrows():
            cinsiyet = temizle_metin(row["CİNSİYET"])
            yas = row["YAŞ"]

            norm_satirlari = norm[
                (norm["CİNSİYET"].astype(str).str.upper().str.strip() == cinsiyet)
                & (norm["YAŞ"] == yas)
            ]

            if norm_satirlari.empty:
                puanlar.append(4)
            else:
                puanlar.append(norm_puanla(row[test], norm_satirlari.iloc[0]))

        ham[puan_sutunu] = puanlar

    for brans, bilgi in BRANS_KRITERLERI.items():
        sonuclar = []
        sira_puanlari = []

        for _, row in ham.iterrows():
            yas = row["YAŞ"]
            cinsiyet = temizle_metin(row["CİNSİYET"])

            if yas not in bilgi["yas"] or cinsiyet not in bilgi["cinsiyet"]:
                sonuclar.append("REFERANS DIŞI")
                sira_puanlari.append(0)
                continue

            basarili = 0
            kalite = 0

            for test in bilgi["kriterler"]:
                puan = row[f"{test} PUAN"]

                if puan >= 12:
                    basarili += 1

                kalite += puan

            sonuclar.append(f"{basarili}/5 - {kalite}")
            sira_puanlari.append(basarili * 1000 + kalite)

        ham[f"{brans} SONUÇ"] = sonuclar
        ham[f"{brans} SIRA"] = sira_puanlari

    branslar = list(BRANS_KRITERLERI.keys())
    onerilenler = []

    for _, row in ham.iterrows():
        en_yuksek = max(row[f"{b} SIRA"] for b in branslar)

        if en_yuksek == 0:
            onerilenler.append("REFERANS DIŞI")
        else:
            secilenler = [
                b for b in branslar
                if row[f"{b} SIRA"] == en_yuksek
            ]
            onerilenler.append(", ".join(secilenler))

    ham["ÖNERİLEN BRANŞ"] = onerilenler
    return ham


def excel_olustur(ham):
    temel_sutunlar = [
        "S.N.",
        "KURUM",
        "BÖLGE",
        "İLÇE",
        "ANTRENÖR ADI",
        "ÜYE NO",
        "AD SOYAD",
        "OKUL",
        "TC KİMLİK",
        "DOĞUM\nTARİHİ",
        "DOĞUM\nYILI",
        "CİNSİYET",
        "VELİ TELEFON 1",
    ]

    test_ve_puan_sutunlari = []

    for test in CIKTI_TEST_SIRASI:
        if test in ham.columns and test not in test_ve_puan_sutunlari:
            test_ve_puan_sutunlari.append(test)

        puan_sutunu = f"{test} PUAN"

        if puan_sutunu in ham.columns and puan_sutunu not in test_ve_puan_sutunlari:
            test_ve_puan_sutunlari.append(puan_sutunu)

    brans_sutunlari = [
        "KARATE SONUÇ",
        "TEKVANDO SONUÇ",
        "BOKS SONUÇ",
        "JUDO SONUÇ",
        "GÜREŞ SONUÇ",
        "ÖNERİLEN BRANŞ",
    ]

    kolonlar = [
        c for c in temel_sutunlar + test_ve_puan_sutunlari + brans_sutunlari
        if c in ham.columns
    ]

    temiz = ham[kolonlar].copy()

    output = BytesIO()
    temiz.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    basliklar = [cell.value for cell in ws[1]]

    for test in CIKTI_TEST_SIRASI:
        puan_sutunu = f"{test} PUAN"

        if test not in basliklar or puan_sutunu not in ham.columns:
            continue

        col_no = basliklar.index(test) + 1

        for i in range(len(temiz)):
            puan = ham.iloc[i][puan_sutunu]

            if puan in RENKLER:
                fill = PatternFill(
                    start_color=RENKLER[puan],
                    end_color=RENKLER[puan],
                    fill_type="solid",
                )

                ws.cell(row=i + 2, column=col_no).fill = fill

    ince_kenar = Side(style="thin", color="000000")

    tam_kenarlik = Border(
        left=ince_kenar,
        right=ince_kenar,
        top=ince_kenar,
        bottom=ince_kenar,
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = tam_kenarlik

    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 30

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                cell_value = str(cell.value)
                longest_line = max(
                    len(line)
                    for line in cell_value.split("\n")
                )
                max_length = max(max_length, longest_line)

        adjusted_width = max_length + 3
        adjusted_width = max(10, min(adjusted_width, 35))
        ws.column_dimensions[column_letter].width = adjusted_width

    final = BytesIO()
    wb.save(final)
    final.seek(0)

    return final


def _norm_dosyasi_bul():
    adaylar = [
        Path(__file__).parent / "NORM TABLO.xlsx",
        Path(__file__).parent / "NORM TABLO(5).xlsx",
        Path.home() / "Desktop" / "NORM TABLO.xlsx",
        Path.home() / "Desktop" / "NORM TABLO(5).xlsx",
    ]
    for yol in adaylar:
        if yol.exists():
            return yol
    return None


def _satir_degeri(row, adaylar, varsayilan=""):
    for aday in adaylar:
        if aday in row and pd.notna(row[aday]):
            return row[aday]
    return varsayilan


def _dolu_deger_mi(deger):
    if pd.isna(deger):
        return False
    return str(deger).strip() != ""


def brans_supabase_ham_verisi(sporcular, testler):
    birlesik = testler.merge(
        sporcular,
        left_on="sporcu_id",
        right_on="id",
        how="left",
        suffixes=("_test", "_sporcu"),
    )

    satirlar = []

    for sira, (_, row) in enumerate(birlesik.iterrows(), start=1):
        dogum_yili = _satir_degeri(row, ["yas", "yas_sporcu", "dogum_yili", "doğum_yılı", "dogum yılı"])
        dogum_tarihi = _satir_degeri(row, ["dogum_tarihi", "doğum_tarihi", "doğum tarihi"])

        ham_satiri = {
            "S.N.": sira,
            "KURUM": _satir_degeri(row, ["kurum"]),
            "BÖLGE": _satir_degeri(row, ["bolge", "bölge"]),
            "İLÇE": _satir_degeri(row, ["ilce", "ilçe"]),
            "ANTRENÖR ADI": _satir_degeri(row, ["antrenor_adi", "antrenör_adı", "antrenör adi"]),
            "ÜYE NO": _satir_degeri(row, ["uye_no", "üye_no", "uye no", "id_sporcu"]),
            "AD SOYAD": _satir_degeri(row, ["ad_soyad", "ad soyad"]),
            "OKUL": _satir_degeri(row, ["okul"]),
            "TC KİMLİK": _satir_degeri(row, ["tc_kimlik", "tc kimlik"]),
            "CİNSİYET": _satir_degeri(row, ["cinsiyet", "cinsiyet_sporcu", "cinsiyet_test"]),
            "VELİ TELEFON 1": _satir_degeri(row, ["veli_telefon_1", "veli telefon 1", "telefon"]),
            "BOY": _satir_degeri(row, ["boy"]),
            "KİLO": _satir_degeri(row, ["kilo"]),
            "BACAK BOYU": _satir_degeri(row, ["bacak_boyu", "bacak boyu"]),
            "OTURMA YÜKSEKLİĞİ": _satir_degeri(row, ["oturma_yuksekligi", "oturma_yüksekliği"]),
            "KULAÇ": _satir_degeri(row, ["kulac", "kulaç"]),
            "FONKSİYONEL ÇÖMELME": _satir_degeri(row, ["fonksiyonel_comelme", "fonksiyonel_çömelme"]),
            "EL ÇABUKLUĞU": _satir_degeri(row, ["el_cabuklugu", "el_çabukluğu"]),
            "AYAK ÇABUKLUĞU": _satir_degeri(row, ["ayak_cabuklugu", "ayak_çabukluğu"]),
            "EL KAVRAMA": _satir_degeri(row, ["el_kavrama"]),
            "SIRT BACAK": _satir_degeri(row, ["sirt_bacak", "sırt_bacak"]),
            "HEXAGON": _satir_degeri(row, ["hexagon"]),
            "DURARAK UZUN ATLAMA": _satir_degeri(row, ["durarak_uzun_atlama"]),
            "GERİYE SAĞLIK TOPU ATMA": _satir_degeri(row, ["geriye_saglik_topu", "geriye_sağlık_topu"]),
            "DİKEY SIÇRAMA": _satir_degeri(row, ["dikey_sicrama", "dikey_sıçrama"]),
            "LANE AGILITY": _satir_degeri(row, ["lane_ceviklik", "lane_çeviklik"]),
            "20 M. SPRINT": _satir_degeri(row, ["sprint20", "20m_sprint", "20_m_sprint"]),
        }

        if _dolu_deger_mi(dogum_yili):
            ham_satiri["DOĞUM\nYILI"] = dogum_yili
        elif _dolu_deger_mi(dogum_tarihi):
            ham_satiri["DOĞUM\nTARİHİ"] = dogum_tarihi

        satirlar.append(ham_satiri)

    return pd.DataFrame(satirlar)


ON_TEST_RENKLER = {
    "ÇOK ALTI": "FF0000",
    "ALTI": "FFC000",
    "ORTALAMA": "5B9BD5",
    "ÜSTÜ": "A9D08E",
    "ÇOK ÜSTÜ": "00B050",
}

ON_TEST_PUANLARI = {
    "ÇOK ALTI": 5,
    "ALTI": 10,
    "ORTALAMA": 15,
    "ÜSTÜ": 20,
    "ÇOK ÜSTÜ": 25,
}

ON_TEST_SEVIYE_SIRASI = list(ON_TEST_PUANLARI.keys())

ON_TEST_TERS_SEVIYE = {
    "ÇOK ALTI": "ÇOK ÜSTÜ",
    "ALTI": "ÜSTÜ",
    "ORTALAMA": "ORTALAMA",
    "ÜSTÜ": "ALTI",
    "ÇOK ÜSTÜ": "ÇOK ALTI",
}

ON_TEST_BARAJLARI = {
    7: 50,
    8: 60,
    9: 70,
    10: 70,
}

ON_TESTLER = {
    "BOY": {
        "sayfa": "BOY",
        "adaylar": ["BOY"],
        "puanlanir": False,
    },
    "KULAÇ": {
        "sayfa": "KULAÇ",
        "adaylar": ["KULAÇ", "KULAC"],
        "puanlanir": False,
    },
    "OTUR UZAN": {
        "sayfa": "OTUR UZAN",
        "adaylar": ["OTUR UZAN", "ESNEKLİK", "ESNEKLIK"],
        "puanlanir": True,
    },
    "DURARAK UZUN ATLAMA": {
        "sayfa": "DURARAK UZUN ATLAMA",
        "adaylar": ["DURARAK UZUN ATLAMA", "UZUN ATLAMA"],
        "puanlanir": True,
    },
    "MEKİK": {
        "sayfa": "MEKİK",
        "adaylar": ["MEKİK", "MEKIK"],
        "puanlanir": True,
    },
    "20 M. SPRINT": {
        "sayfa": "20 M. SPRINT",
        "adaylar": ["20 M. SPRINT", "20M SPRINT", "20 M SPRINT", "SPRINT"],
        "puanlanir": True,
        "ters_mantik": True,
    },
}


def on_test_normalize_text(x):
    x = str(x).strip().upper()
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = x.replace("\n", " ")
    x = re.sub(r"\s+", " ", x)
    return x


def on_test_temizle_metin(x):
    return str(x).strip().upper()


def on_test_sayiya_cevir(x):
    return float(str(x).replace(",", ".").strip())


def on_test_kolon_bul(df, adaylar):
    normalized_cols = {on_test_normalize_text(c): c for c in df.columns}

    for aday in adaylar:
        aday_norm = on_test_normalize_text(aday)

        for norm_col, original_col in normalized_cols.items():
            if aday_norm in norm_col:
                return original_col

    return None


def on_test_tam_kolon_bul(df, aday):
    normalized_cols = {on_test_normalize_text(c): c for c in df.columns}
    aday_norm = on_test_normalize_text(aday)

    if aday_norm in normalized_cols:
        return normalized_cols[aday_norm]

    return on_test_kolon_bul(df, [aday])


def on_test_aralik_uyuyor_mu(deger, aralik):
    metin = str(aralik).replace(",", ".").strip()
    sayilar = re.findall(r"\d+(?:\.\d+)?", metin)

    if not sayilar:
        return False

    sayilar = [float(s) for s in sayilar]

    if "≤" in metin or "<=" in metin:
        return deger <= sayilar[0]

    if "≥" in metin or ">=" in metin:
        return deger >= sayilar[0]

    if "-" in metin and len(sayilar) >= 2:
        alt = min(sayilar[0], sayilar[1])
        ust = max(sayilar[0], sayilar[1])
        return alt <= deger <= ust

    return False


def on_test_aralik_ilk_sayi(aralik):
    metin = str(aralik).replace(",", ".").strip()
    sayilar = re.findall(r"\d+(?:\.\d+)?", metin)

    if not sayilar:
        return None

    return float(sayilar[0])


def on_test_norm_dosyasi_bul():
    adaylar = [
        Path(__file__).parent / "NORM TABLO ÖN TESTLER.xlsx",
        Path(__file__).parent / "NORM TABLO ON TESTLER.xlsx",
        Path.home() / "Desktop" / "NORM TABLO ÖN TESTLER.xlsx",
        Path.home() / "Desktop" / "NORM TABLO ON TESTLER.xlsx",
    ]
    for yol in adaylar:
        if yol.exists():
            return yol

    for klasor in [Path(__file__).parent, Path.home() / "Desktop"]:
        if not klasor.exists():
            continue

        for yol in klasor.glob("*.xlsx"):
            ad = on_test_normalize_text(yol.name)
            if "NORM" in ad and "ON TEST" in ad:
                return yol

    return None


def on_test_sheet_oku(norm_dosya, sheet_adi):
    excel = pd.ExcelFile(norm_dosya)
    sayfalar = {
        on_test_normalize_text(s): s
        for s in excel.sheet_names
    }

    sayfa_norm = on_test_normalize_text(sheet_adi)
    if sayfa_norm not in sayfalar:
        raise ValueError(
            f"Ön Test norm dosyasında '{sheet_adi}' sayfası bulunamadı. "
            f"Mevcut sayfalar: {excel.sheet_names}"
        )

    return pd.read_excel(norm_dosya, sheet_name=sayfalar[sayfa_norm])


def on_test_norm_indexi_hazirla(norm_dosya):
    excel = pd.ExcelFile(norm_dosya)
    sayfalar = {
        on_test_normalize_text(s): s
        for s in excel.sheet_names
    }
    norm_index = {}

    for test, bilgi in ON_TESTLER.items():
        sayfa_norm = on_test_normalize_text(bilgi["sayfa"])

        if sayfa_norm not in sayfalar:
            raise ValueError(
                f"Ön Test norm dosyasında '{bilgi['sayfa']}' sayfası bulunamadı. "
                f"Mevcut sayfalar: {excel.sheet_names}"
            )

        norm = excel.parse(sheet_name=sayfalar[sayfa_norm])
        cinsiyet_col = on_test_tam_kolon_bul(norm, "CİNSİYET")
        yas_col = on_test_tam_kolon_bul(norm, "YAŞ")

        if cinsiyet_col is None or yas_col is None:
            raise ValueError(f"'{bilgi['sayfa']}' norm sayfasında CİNSİYET veya YAŞ sütunu bulunamadı.")

        seviye_kolonlari = {
            seviye: on_test_tam_kolon_bul(norm, seviye)
            for seviye in ON_TEST_SEVIYE_SIRASI
        }

        for _, row in norm.iterrows():
            yas = pd.to_numeric(row[yas_col], errors="coerce")

            if pd.isna(yas):
                continue

            kurallar = {
                seviye: row[kolon] if kolon is not None and kolon in row else None
                for seviye, kolon in seviye_kolonlari.items()
            }
            anahtar = (
                test,
                on_test_temizle_metin(row[cinsiyet_col]),
                int(yas),
            )
            norm_index[anahtar] = kurallar

    return norm_index


def on_test_yas_hesapla(ham):
    dogum_tarihi_col = on_test_kolon_bul(ham, ["DOĞUM TARİHİ", "DOGUM TARIHI"])
    dogum_yili_col = on_test_kolon_bul(ham, ["DOĞUM YILI", "DOGUM YILI"])
    yas_col = on_test_kolon_bul(ham, ["YAŞ", "YAS"])
    bu_yil = pd.Timestamp.now().year
    yas = None

    if dogum_tarihi_col:
        yas = bu_yil - pd.to_datetime(ham[dogum_tarihi_col], errors="coerce").dt.year

    if dogum_yili_col:
        dogum_yili_yasi = bu_yil - pd.to_numeric(ham[dogum_yili_col], errors="coerce")
        yas = dogum_yili_yasi if yas is None else yas.fillna(dogum_yili_yasi)

    if yas_col:
        hazir_yas = pd.to_numeric(ham[yas_col], errors="coerce")
        yas = hazir_yas if yas is None else yas.fillna(hazir_yas)

    if yas is None:
        yas = pd.Series([0] * len(ham), index=ham.index)

    ham["YAŞ"] = yas.fillna(0).astype(int)
    return ham


def on_test_veriyi_standartlastir(ham):
    ham = ham.copy()

    cinsiyet_col = on_test_kolon_bul(ham, ["CİNSİYET", "CINSIYET"])
    if cinsiyet_col and cinsiyet_col != "CİNSİYET":
        ham["CİNSİYET"] = ham[cinsiyet_col]

    for test, bilgi in ON_TESTLER.items():
        kolon = on_test_kolon_bul(ham, bilgi["adaylar"])
        if kolon and kolon != test:
            ham[test] = ham[kolon]

    return ham


def on_test_norm_satiri_ters_mi(norm_kurallari):
    cok_alti = on_test_aralik_ilk_sayi(norm_kurallari.get("ÇOK ALTI"))
    cok_ustu = on_test_aralik_ilk_sayi(norm_kurallari.get("ÇOK ÜSTÜ"))

    if cok_alti is None or cok_ustu is None:
        return False

    return cok_alti > cok_ustu


def on_test_norm_seviye_bul(deger, norm_kurallari, ters_mantik=False):
    if pd.isna(deger) or on_test_temizle_metin(deger) in ["VERİ YOK", "VERI YOK", "NAN", ""]:
        return "ÇOK ALTI"

    try:
        deger = on_test_sayiya_cevir(deger)
    except Exception:
        return "ÇOK ALTI"

    seviye = "ÇOK ALTI"

    for kolon in ON_TEST_SEVIYE_SIRASI:
        if kolon in norm_kurallari and on_test_aralik_uyuyor_mu(deger, norm_kurallari[kolon]):
            seviye = kolon
            break

    if ters_mantik and not on_test_norm_satiri_ters_mi(norm_kurallari):
        return ON_TEST_TERS_SEVIYE[seviye]

    return seviye


def on_test_kriter_hesapla(ham, puan_sutunlari):
    ham["TOPLAM ÖN TEST PUANI"] = ham[puan_sutunlari].sum(axis=1)
    barajlar = ham["YAŞ"].map(ON_TEST_BARAJLARI)
    cagrilacak = barajlar.notna() & (ham["TOPLAM ÖN TEST PUANI"] >= barajlar)
    ham["EUROFIT DURUMU"] = "ÇAĞRILMAYACAK"
    ham.loc[cagrilacak, "EUROFIT DURUMU"] = "ÇAĞRILACAK"
    return ham


def on_test_degerlendir(ham, norm_dosya):
    ham = on_test_veriyi_standartlastir(ham)
    ham = on_test_yas_hesapla(ham)
    norm_index = on_test_norm_indexi_hazirla(norm_dosya)

    if "CİNSİYET" not in ham.columns:
        ham["CİNSİYET"] = ""

    temiz_cinsiyetler = ham["CİNSİYET"].map(on_test_temizle_metin)
    yaslar = pd.to_numeric(ham["YAŞ"], errors="coerce").fillna(0).astype(int)
    puan_sutunlari = []

    for test, bilgi in ON_TESTLER.items():
        seviye_sutunu = f"{test} SEVİYE"
        puan_sutunu = f"{test} PUAN"

        if bilgi["puanlanir"]:
            puan_sutunlari.append(puan_sutunu)

        if test not in ham.columns:
            ham[test] = "VERİ YOK"
            ham[seviye_sutunu] = "ÇOK ALTI"
            if bilgi["puanlanir"]:
                ham[puan_sutunu] = ON_TEST_PUANLARI["ÇOK ALTI"]
            else:
                ham[f"{test} AÇIKLAMA"] = ham[seviye_sutunu]
            continue

        seviyeler = []
        degerler = ham[test]

        for deger, cinsiyet, yas in zip(degerler, temiz_cinsiyetler, yaslar):
            norm_kurallari = norm_index.get((test, cinsiyet, yas))

            if norm_kurallari is None:
                seviye = "ÇOK ALTI"
            else:
                seviye = on_test_norm_seviye_bul(
                    deger,
                    norm_kurallari,
                    ters_mantik=bilgi.get("ters_mantik", False),
                )

            seviyeler.append(seviye)

        ham[seviye_sutunu] = seviyeler
        if bilgi["puanlanir"]:
            ham[puan_sutunu] = ham[seviye_sutunu].map(ON_TEST_PUANLARI)
        else:
            ham[f"{test} AÇIKLAMA"] = ham[seviye_sutunu]

    ham = on_test_kriter_hesapla(ham, puan_sutunlari)
    return ham


def on_test_excel_olustur(ham):
    kimlik_sutunlari = [
        "S.N.",
        "AD SOYAD",
        "OKUL",
        "İLÇE",
        "DOĞUM\nTARİHİ",
        "DOĞUM\nYILI",
        "CİNSİYET",
        "YAŞ",
    ]

    test_sutunlari = []
    for test, bilgi in ON_TESTLER.items():
        if test in ham.columns:
            test_sutunlari.append(test)

        aciklama_sutunu = f"{test} AÇIKLAMA"
        if not bilgi["puanlanir"] and aciklama_sutunu in ham.columns:
            test_sutunlari.append(aciklama_sutunu)

        puan_sutunu = f"{test} PUAN"
        if bilgi["puanlanir"] and puan_sutunu in ham.columns:
            test_sutunlari.append(puan_sutunu)

    sonuc_sutunlari = [
        "TOPLAM ÖN TEST PUANI",
        "EUROFIT DURUMU",
    ]

    kolonlar = [
        c for c in kimlik_sutunlari + test_sutunlari + sonuc_sutunlari
        if c in ham.columns
    ]

    temiz = ham[kolonlar].copy()

    output = BytesIO()
    temiz.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    basliklar = [cell.value for cell in ws[1]]
    baslik_index = {
        baslik: sira + 1
        for sira, baslik in enumerate(basliklar)
    }
    seviye_fills = {
        seviye: PatternFill(
            start_color=renk,
            end_color=renk,
            fill_type="solid",
        )
        for seviye, renk in ON_TEST_RENKLER.items()
    }

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for test in ON_TESTLER:
        seviye_sutunu = f"{test} SEVİYE"

        if test not in baslik_index or seviye_sutunu not in ham.columns:
            continue

        test_col_no = baslik_index[test]
        seviyeler = ham[seviye_sutunu].tolist()

        for row_offset, seviye in enumerate(seviyeler, start=2):
            if row_offset > ws.max_row:
                break

            if seviye in seviye_fills:
                ws.cell(row=row_offset, column=test_col_no).fill = seviye_fills[seviye]

    ws.row_dimensions[1].height = 30

    for column_cells in ws.columns:
        baslik = str(column_cells[0].value or "")
        column_letter = get_column_letter(column_cells[0].column)
        adjusted_width = max(len(baslik) + 3, 10)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 35)

    final = BytesIO()
    wb.save(final)
    final.seek(0)

    return final


SAYFALAR = [
    "🏠 Ana Sayfa",
    "🧒 Sporcu Kayıt",
    "📋 Test Veri Girişi",
    "📊 Sonuçlar",
    "📈 Dashboard",
    "🧪 Ön Testler",
    "🇪🇺 Eurofit",
    "🏅 Branş Amaçlı",
]

with st.sidebar:
    logo_yolu = Path(__file__).parent / "logo.png"

    if logo_yolu.exists():
        st.image(str(logo_yolu), width=180)

    st.title("USGEP Menü")
    sayfa = st.radio("Sayfa seç", SAYFALAR, label_visibility="collapsed")

test_modu = st.query_params.get("test", "normal")


# ---------------------------
# ANA SAYFA
# ---------------------------

if sayfa == "🏠 Ana Sayfa":

    st.title("USGEP Branş Yönlendirme Sistemi")

    st.markdown("""
    ### Sistem Durumu

    ✅ Online erişim aktif  
    ✅ Tablet erişimi aktif  
    ✅ Streamlit Cloud yayında  
    ✅ Supabase veritabanı bağlı  
    ✅ QR istasyon sistemi aktif  
    ✅ Admin yetkili düzenleme aktif  

    ---
    """)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Tablet", "6")

    with col2:
        st.metric("Branş", "5")

    with col3:
        st.metric("Test Alanı", "17")


# ---------------------------
# SPORCU KAYIT
# ---------------------------

elif sayfa == "🧒 Sporcu Kayıt":

    st.title("Sporcu Kayıt")

    st.subheader("Tekli Sporcu Kayıt")

    with st.form("tekli_sporcu_kayit_formu"):

        ad = st.text_input("Ad Soyad")
        yas = st.number_input("Doğum Yılı", 2000, 2025)
        cinsiyet = st.selectbox("Cinsiyet", ["ERKEK", "KIZ"])
        ilce = st.text_input("İlçe")

        submit = st.form_submit_button("Kaydet")

        if submit:
            if not ad.strip():
                st.warning("Ad soyad boş bırakılamaz.")
            else:
                supabase.table("sporcular").insert({
                    "ad_soyad": ad.strip(),
                    "yas": int(yas),
                    "cinsiyet": cinsiyet,
                    "ilce": ilce.strip()
                }).execute()

                st.success("Sporcu kaydedildi.")

    st.divider()

    st.subheader("Excel ile Toplu Sporcu Yükleme")

    st.info("Excel sütunları şu şekilde olmalı: ad_soyad | yas | cinsiyet | ilce")

    dosya = st.file_uploader(
        "Excel Dosyası Yükle",
        type=["xlsx"]
    )

    if dosya is not None:

        try:
            df = pd.read_excel(dosya)

            df.columns = (
                df.columns
                .astype(str)
                .str.strip()
                .str.lower()
            )

            st.write("Önizleme")
            st.dataframe(df.head(), use_container_width=True)

            gerekli_kolonlar = [
                "ad_soyad",
                "yas",
                "cinsiyet",
                "ilce"
            ]

            eksik = [
                kolon for kolon in gerekli_kolonlar
                if kolon not in df.columns
            ]

            if eksik:
                st.error(f"Eksik sütunlar: {eksik}")

            else:
                if st.button("Toplu Yüklemeyi Başlat"):

                    df = df[gerekli_kolonlar].copy()

                    df["ad_soyad"] = df["ad_soyad"].astype(str).str.strip()
                    df["yas"] = df["yas"].astype(int)
                    df["cinsiyet"] = df["cinsiyet"].astype(str).str.strip().str.upper()
                    df["ilce"] = df["ilce"].astype(str).str.strip()

                    df = df[df["ad_soyad"] != ""]

                    veriler = df.to_dict(orient="records")

                    if len(veriler) == 0:
                        st.warning("Yüklenecek geçerli sporcu bulunamadı.")
                    else:
                        supabase.table("sporcular").insert(veriler).execute()
                        st.success(f"{len(veriler)} sporcu başarıyla yüklendi.")

        except Exception as e:
            st.error(f"Hata oluştu: {e}")


# ---------------------------
# NORMAL TEST VERİ GİRİŞİ
# ---------------------------

elif sayfa == "📋 Test Veri Girişi":

    st.title("Test Veri Girişi")

    sporcular = sporculari_getir()

    if sporcular.empty:
        st.warning("Önce sporcu kaydı oluşturun.")
    else:
        sporcular["secim"] = sporcular["id"].astype(str) + " - " + sporcular["ad_soyad"]

        secili = st.selectbox("Sporcu Seç", sporcular["secim"])
        sporcu_id = int(secili.split(" - ")[0])

        st.subheader("Testler")

        boy = st.number_input("Boy", min_value=0.0, step=0.1)
        kilo = st.number_input("Kilo", min_value=0.0, step=0.1)
        sprint20 = st.number_input("20m Sprint", min_value=0.0, step=0.01)
        dikey_sicrama = st.number_input("Dikey Sıçrama", min_value=0.0, step=0.1)

        if st.button("Testleri Kaydet"):
            mevcut = supabase.table("testler").select("*").eq("sporcu_id", sporcu_id).execute()

            veri = {
                "sporcu_id": sporcu_id,
                "boy": float(boy),
                "kilo": float(kilo),
                "sprint20": float(sprint20),
                "dikey_sicrama": float(dikey_sicrama)
            }

            if mevcut.data:
                supabase.table("testler").update(veri).eq("sporcu_id", sporcu_id).execute()
            else:
                supabase.table("testler").insert(veri).execute()

            st.success("Test verileri kaydedildi.")


# ---------------------------
# SONUÇLAR
# ---------------------------

elif sayfa == "📊 Sonuçlar":

    st.title("Sonuçlar")

    sporcular = sporculari_getir()
    testler = testleri_getir()

    if testler.empty:
        st.info("Henüz test sonucu bulunmuyor.")
    else:
        if not sporcular.empty:
            sonuc = testler.merge(
                sporcular,
                left_on="sporcu_id",
                right_on="id",
                how="left",
                suffixes=("_test", "_sporcu")
            )
        else:
            sonuc = testler

        st.dataframe(sonuc, use_container_width=True)


# ---------------------------
# DASHBOARD
# ---------------------------

elif sayfa == "📈 Dashboard":

    st.title("Dashboard")

    sporcular = sporculari_getir()
    testler = testleri_getir()

    col1, col2 = st.columns(2)

    with col1:
        st.metric("Toplam Sporcu", len(sporcular))

    with col2:
        st.metric("Toplam Test Kaydı", len(testler))

    st.subheader("Sporcu Listesi")
    st.dataframe(sporcular, use_container_width=True)

    st.subheader("Test Listesi")
    st.dataframe(testler, use_container_width=True)

# ---------------------------
# ÖN TESTLER
# ---------------------------

elif sayfa == "🧪 Ön Testler":

    st.title("Ön Testler")

    st.info("Ön Testler, Branş Amaçlı sistemden bağımsız ayrı norm ve kriter dosyasıyla değerlendirilir.")

    col1, col2 = st.columns(2)

    with col1:
        on_test_ham_dosya = st.file_uploader(
            "Ön test ham Excel dosyasını yükle",
            type=["xlsx"],
            key="on_test_ham_dosya",
        )

    with col2:
        on_test_yuklenen_norm = st.file_uploader(
            "Ön test norm Excel dosyasını yükle",
            type=["xlsx"],
            key="on_test_norm_dosya",
            help="TEST | CİNSİYET | YAŞ | ÇOK ALTI | ALTI | ORTALAMA | ÜSTÜ | ÇOK ÜSTÜ yapısında olmalı.",
        )

    if on_test_ham_dosya is None:
        st.warning("Ön Testler için ham Excel dosyasını yükleyin.")
    elif on_test_yuklenen_norm is None:
        st.warning("Ön Testler için norm Excel dosyasını yükleyin.")
    else:
        try:
            on_test_ham = pd.read_excel(on_test_ham_dosya)
            on_test_source_id = (
                on_test_ham_dosya.name,
                getattr(on_test_ham_dosya, "size", None),
                on_test_yuklenen_norm.name,
                getattr(on_test_yuklenen_norm, "size", None),
            )

            if st.session_state.get("on_test_source_id") != on_test_source_id:
                st.session_state.pop("on_test_sonuc", None)
                st.session_state.pop("on_test_excel_tumu", None)
                st.session_state.pop("on_test_excel_cagrilacak", None)
                st.session_state["on_test_source_id"] = on_test_source_id

            metrik1, metrik2, metrik3 = st.columns(3)

            with metrik1:
                st.metric("Yüklenen Öğrenci", len(on_test_ham))

            with metrik2:
                st.metric("Ön Test", len(ON_TESTLER))

            with metrik3:
                st.metric("Sistem", "Bağımsız")

            st.subheader("Ön Test Ham Veri Önizleme")
            st.dataframe(on_test_ham.head(), use_container_width=True)

            if st.button("Ön Testleri Değerlendir"):
                with st.spinner("Ön testler değerlendiriliyor, lütfen bekleyin..."):
                    on_test_sonuc = on_test_degerlendir(
                        on_test_ham.copy(),
                        on_test_yuklenen_norm,
                    )

                st.session_state["on_test_sonuc"] = on_test_sonuc
                st.session_state.pop("on_test_excel_tumu", None)
                st.session_state.pop("on_test_excel_cagrilacak", None)

            if "on_test_sonuc" in st.session_state:
                on_test_sonuc = st.session_state["on_test_sonuc"]
                cagrilacak = on_test_sonuc[
                    on_test_sonuc["EUROFIT DURUMU"] == "ÇAĞRILACAK"
                ].copy()
                cagrilmayacak_sayi = len(on_test_sonuc) - len(cagrilacak)

                ozet1, ozet2, ozet3 = st.columns(3)

                with ozet1:
                    st.metric("Toplam Öğrenci", len(on_test_sonuc))

                with ozet2:
                    st.metric("Çağrılacak Öğrenci", len(cagrilacak))

                with ozet3:
                    st.metric("Çağrılmayacak Öğrenci", cagrilmayacak_sayi)

                st.subheader("Ön Test Sonuç Önizleme")

                on_test_onizleme = [
                    "S.N.",
                    "AD SOYAD",
                    "CİNSİYET",
                    "YAŞ",
                    "BOY",
                    "BOY AÇIKLAMA",
                    "KULAÇ",
                    "KULAÇ AÇIKLAMA",
                    "OTUR UZAN",
                    "OTUR UZAN PUAN",
                    "DURARAK UZUN ATLAMA",
                    "DURARAK UZUN ATLAMA PUAN",
                    "MEKİK",
                    "MEKİK PUAN",
                    "20 M. SPRINT",
                    "20 M. SPRINT PUAN",
                    "TOPLAM ÖN TEST PUANI",
                    "EUROFIT DURUMU",
                ]

                st.dataframe(
                    on_test_sonuc[[c for c in on_test_onizleme if c in on_test_sonuc.columns]],
                    use_container_width=True,
                    hide_index=True,
                )

                indirme1, indirme2 = st.columns(2)

                with indirme1:
                    if st.button("Tüm Öğrenciler Excel Hazırla", key="on_test_excel_tumu_hazirla"):
                        with st.spinner("Tüm öğrenciler Excel çıktısı hazırlanıyor..."):
                            st.session_state["on_test_excel_tumu"] = on_test_excel_olustur(
                                on_test_sonuc
                            ).getvalue()

                    if "on_test_excel_tumu" in st.session_state:
                        st.download_button(
                            label="Tüm Öğrenciler Excel İndir",
                            data=st.session_state["on_test_excel_tumu"],
                            file_name="USGEP_ON_TEST_TUM_OGRENCILER.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                with indirme2:
                    if st.button("Eurofit’e Çağrılacaklar Excel Hazırla", key="on_test_excel_cagrilacak_hazirla"):
                        with st.spinner("Eurofit’e çağrılacaklar Excel çıktısı hazırlanıyor..."):
                            st.session_state["on_test_excel_cagrilacak"] = on_test_excel_olustur(
                                cagrilacak
                            ).getvalue()

                    if "on_test_excel_cagrilacak" in st.session_state:
                        st.download_button(
                            label="Eurofit’e Çağrılacaklar Excel İndir",
                            data=st.session_state["on_test_excel_cagrilacak"],
                            file_name="USGEP_ON_TEST_EUROFIT_CAGRILACAKLAR.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

        except Exception as e:
            st.error(f"Ön Test değerlendirmesi oluşturulamadı: {e}")


# ---------------------------
# EUROFIT
# ---------------------------

elif sayfa == "🇪🇺 Eurofit":

    st.title("Eurofit")

    st.info("Bu sayfada Eurofit testleri için değerlendirme, puanlama ve renklendirme yapılacak.")

    sporcular = sporculari_getir()
    testler = testleri_getir()

    st.subheader("Sporcu Verileri")
    st.dataframe(sporcular, use_container_width=True)

    st.subheader("Test Verileri")
    st.dataframe(testler, use_container_width=True)


# ---------------------------
# BRANŞ AMAÇLI
# ---------------------------

elif sayfa == "🏅 Branş Amaçlı":

    st.title("Branş Amaçlı")

    sporcular = sporculari_getir()
    testler = testleri_getir()

    if sporcular.empty:
        st.warning("Önce sporcu kaydı oluşturun.")
    elif testler.empty:
        st.warning("Henüz test sonucu bulunmuyor.")
    else:
        norm_dosyasi = _norm_dosyasi_bul()
        yuklenen_norm = st.file_uploader(
            "Norm tablo Excel dosyasını yükle",
            type=["xlsx"],
            help="NORM TABLO.xlsx bulunamazsa buradan yükleyebilirsiniz.",
        )

        kaynak = yuklenen_norm if yuklenen_norm is not None else norm_dosyasi

        if kaynak is None:
            st.error("NORM TABLO.xlsx bulunamadı. Norm dosyasını bu sayfadan yükleyin.")
        else:
            ham = brans_supabase_ham_verisi(sporcular, testler)

            metrik1, metrik2, metrik3 = st.columns(3)

            with metrik1:
                st.metric("Yüklenen Öğrenci", len(ham))

            with metrik2:
                st.metric("Branş", "5")

            with metrik3:
                st.metric("Test / Ölçüm Alanı", "17")

            st.subheader("Ham Veri Önizleme")
            st.dataframe(ham.head(), use_container_width=True)

            if st.button("Puanla ve Branşları Hesapla"):
                try:
                    sonuc = islem_yap(ham.copy(), kaynak)

                    st.write("İşlenen sonuç öğrenci sayısı:", len(sonuc))

                    st.subheader("Sonuç Önizleme")
                    onizleme_kolonlari = [
                        "AD SOYAD",
                        "KARATE SONUÇ",
                        "TEKVANDO SONUÇ",
                        "BOKS SONUÇ",
                        "JUDO SONUÇ",
                        "GÜREŞ SONUÇ",
                        "ÖNERİLEN BRANŞ",
                    ]
                    st.dataframe(
                        sonuc[[c for c in onizleme_kolonlari if c in sonuc.columns]].head(30),
                        use_container_width=True,
                        hide_index=True,
                    )

                    excel_dosya = excel_olustur(sonuc)

                    st.download_button(
                        label="Sonuç Excel Dosyasını İndir",
                        data=excel_dosya,
                        file_name="USGEP_SONUCLAR.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                except Exception as e:
                    st.error(str(e))


# ---------------------------
# QR / TEST İSTASYON MODU
# ---------------------------

TEST_ISTASYONLARI = {

    "boy": {
        "baslik": "Boy Ölçüm İstasyonu",
        "kolon": "boy",
        "etiket": "Boy",
        "birim": "cm",
        "step": 0.1
    },

    "kilo": {
        "baslik": "Kilo Ölçüm İstasyonu",
        "kolon": "kilo",
        "etiket": "Kilo",
        "birim": "kg",
        "step": 0.1
    },

    "kulac": {
        "baslik": "Kulaç Ölçüm İstasyonu",
        "kolon": "kulac",
        "etiket": "Kulaç",
        "birim": "cm",
        "step": 0.1
    },

    "durarak_uzun_atlama": {
        "baslik": "Durarak Uzun Atlama İstasyonu",
        "kolon": "durarak_uzun_atlama",
        "etiket": "Durarak Uzun Atlama",
        "birim": "cm",
        "step": 0.1
    },

    "dikey_sicrama": {
        "baslik": "Dikey Sıçrama İstasyonu",
        "kolon": "dikey_sicrama",
        "etiket": "Dikey Sıçrama",
        "birim": "cm",
        "step": 0.1
    },

    "el_kavrama": {
        "baslik": "El Kavrama İstasyonu",
        "kolon": "el_kavrama",
        "etiket": "El Kavrama",
        "birim": "kg",
        "step": 0.1
    },

    "geriye_saglik_topu": {
        "baslik": "Geriye Sağlık Topu İstasyonu",
        "kolon": "geriye_saglik_topu",
        "etiket": "Geriye Sağlık Topu",
        "birim": "m",
        "step": 0.1
    },

    "sprint": {
        "baslik": "20m Sprint İstasyonu",
        "kolon": "sprint20",
        "etiket": "20m Sprint",
        "birim": "sn",
        "step": 0.01
    },

    "ayak_cabuklugu": {
        "baslik": "Ayak Çabukluğu İstasyonu",
        "kolon": "ayak_cabuklugu",
        "etiket": "Ayak Çabukluğu",
        "birim": "sn",
        "step": 0.01
    },

    "el_cabuklugu": {
        "baslik": "El Çabukluğu İstasyonu",
        "kolon": "el_cabuklugu",
        "etiket": "El Çabukluğu",
        "birim": "adet",
        "step": 1.0
    },

    "sirt_bacak": {
        "baslik": "Sırt Bacak Kuvveti İstasyonu",
        "kolon": "sirt_bacak",
        "etiket": "Sırt Bacak Kuvveti",
        "birim": "kg",
        "step": 0.1
    },

    "hexagon": {
        "baslik": "Hexagon İstasyonu",
        "kolon": "hexagon",
        "etiket": "Hexagon",
        "birim": "sn",
        "step": 0.01
    },

    "lane_ceviklik": {
        "baslik": "Lane Çeviklik İstasyonu",
        "kolon": "lane_ceviklik",
        "etiket": "Lane Çeviklik",
        "birim": "sn",
        "step": 0.01
    }
}


if test_modu in TEST_ISTASYONLARI:

    test = TEST_ISTASYONLARI[test_modu]

    st.title(test["baslik"])

    sporcular = sporculari_getir()

    if sporcular.empty:
        st.warning("Henüz sporcu kaydı yok.")
    else:

        sporcular["secim"] = (
            sporcular["id"].astype(str)
            + " - "
            + sporcular["ad_soyad"]
        )

        secili = st.selectbox(
            "Sporcu",
            sporcular["secim"]
        )

        sporcu_id = int(secili.split(" - ")[0])

        mevcut = supabase.table("testler").select("*").eq("sporcu_id", sporcu_id).execute()

        mevcut_veri = None

        if mevcut.data:
            mevcut_veri = mevcut.data[0].get(test["kolon"])

        admin_sifre = st.sidebar.text_input(
            "Admin Şifresi",
            type="password"
        )

        admin_mi = admin_sifre == ADMIN_PASSWORD

        kilitli = temiz_deger_mi(mevcut_veri) and not admin_mi

        if kilitli:

            st.warning("Bu test sonucu daha önce kaydedilmiş. Düzenleme yetkiniz yok.")

            st.number_input(
                f'{test["etiket"]} ({test["birim"]})',
                value=float(mevcut_veri),
                disabled=True
            )

        else:

            sonuc = st.number_input(
                f'{test["etiket"]} ({test["birim"]})',
                min_value=0.0,
                step=test["step"],
                value=float(mevcut_veri) if temiz_deger_mi(mevcut_veri) else 0.0
            )

            if st.button("Kaydet"):

                if mevcut.data:

                    supabase.table("testler").update({
                        test["kolon"]: float(sonuc)
                    }).eq(
                        "sporcu_id",
                        sporcu_id
                    ).execute()

                else:

                    supabase.table("testler").insert({
                        "sporcu_id": sporcu_id,
                        test["kolon"]: float(sonuc)
                    }).execute()

                st.success(f'{test["etiket"]} kaydedildi.')
                st.rerun()

